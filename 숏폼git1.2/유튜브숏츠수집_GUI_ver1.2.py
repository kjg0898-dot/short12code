# pyinstaller -w -F --add-data "youtube1.2.ui;./" 유튜브숏츠수집_GUI_ver1.2.py

import os
import sys
import time
from datetime import datetime, timedelta
import traceback
import random

from PyQt5.QtWidgets import *
from PyQt5 import uic
from PyQt5.QtGui import *
from PyQt5.QtCore import Qt, QTimer, QTime, QDate
from PyQt5.QtTest import *

from openpyxl import load_workbook
from openpyxl.styles import NamedStyle
from openpyxl.styles import Font, Alignment

from bs4 import BeautifulSoup
from selenium.webdriver.chrome.service import Service
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.alert import Alert
from selenium.common.exceptions import UnexpectedAlertPresentException
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from urllib.parse import urlparse, parse_qs
import re
import urllib.parse

# from youtube_transcript_api import YouTubeTranscriptApi
# from youtube_transcript_api._errors import TranscriptsDisabled, NoTranscriptFound

from yt_dlp import YoutubeDL
import requests
import json


if getattr(sys, 'frozen', False):
    #test.exe로 실행한 경우,test.exe를 보관한 디렉토리의 full path를 취득
    program_directory = os.path.dirname(os.path.abspath(sys.executable))
else:
    #python test.py로 실행한 경우,test.py를 보관한 디렉토리의 full path를 취득
    program_directory = os.path.dirname(os.path.abspath(__file__))

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

UI_PATH = "youtube1.2.ui"

save_date = datetime.today().strftime("%Y%m%d_%H%M")
favicon_image = program_directory + "\\data\\icon-16x16.png"
yj_image = program_directory + "\\data\\icon.png"
exl_sample = program_directory + "\\data\\result_sample.xlsx"
scriptidx_file = program_directory + f"\\data\\scriptidx.ini"
setting_file_1 = program_directory + f"\\data\\setting_1.ini"
setting_file_2 = program_directory + f"\\data\\setting_2.ini"
setting_file_3 = program_directory + f"\\data\\setting_3.ini"


class MainDialog(QDialog):
    def __init__(self):
        QDialog.__init__(self, None)

        self.setWindowFlags(Qt.WindowCloseButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowMinimizeButtonHint)  # 최소화 버튼
        uic.loadUi(os.path.join(BASE_DIR, UI_PATH), self)

        self.setWindowTitle('유튜브 숏츠 정보수집 프로그램 Ver 1.2')
        self.setWindowIcon(QIcon(favicon_image))
        pm = QPixmap(yj_image)
        pm = pm.scaledToWidth(220)
        self.yj_image_btn.setPixmap(pm)

        # 업로드날짜 필터 라디오 버튼 그룹화 (지난 1시간 제거됨)
        self.ud_button_group = QButtonGroup(self)
        self.ud_button_group.addButton(self.udfilter_btn_1)
        self.ud_button_group.addButton(self.udfilter_btn_3)
        self.ud_button_group.addButton(self.udfilter_btn_4)
        self.ud_button_group.addButton(self.udfilter_btn_5)
        self.ud_button_group.addButton(self.udfilter_btn_6)

        self.udfilter_btn_1.setChecked(True)

        # 우선순위 필터 라디오 버튼 그룹화 (선택안함/관련성/인기도)
        self.priority_button_group = QButtonGroup(self)
        self.priority_button_group.addButton(self.priority_btn_0)
        self.priority_button_group.addButton(self.priority_btn_1)
        self.priority_button_group.addButton(self.priority_btn_2)
        self.priority_btn_0.setChecked(True)


        self.layout = QVBoxLayout(self.scrollArea)

        # Scroll Area
        self.scroll_area = QScrollArea()
        self.scroll_area.setWidgetResizable(True)
        self.scroll_widget = QWidget()
        self.scroll_layout = QVBoxLayout(self.scroll_widget)
        self.scroll_area.setWidget(self.scroll_widget)
      
        self.layout.addWidget(self.scroll_area)

        self.rows = []
        for _ in range(100):
            self.add_row()

        self.add_btn.clicked.connect(self.add_row) # 폴더추가 버튼 이벤트
        self.start_btn.clicked.connect(self.main) # 수집시작 버튼 이벤트
        self.exlload_btn.clicked.connect(self.exl_load) # 엑셀 Import 버튼 이벤트
        self.folder_reset_btn.clicked.connect(self.folder_reset) # 폴더명 리셋 버튼 이벤트
        self.folder_sel_btn.clicked.connect(self.folder_path) # 저장 기본폴더 버튼 이벤트
        self.setting_save_btn.clicked.connect(self.setting_save) # 설정저장

        if os.path.exists(setting_file_1):

            with open(setting_file_1, 'r') as file:
                lines = file.readlines()

            # 업로드 날짜 필터 (지난 1시간 제거됨 - 인덱스 조정)
            try:
                if lines[0].strip() == "True" :
                    self.udfilter_btn_1.setChecked(True)
            except :
                pass
            # lines[1]은 이제 사용하지 않음 (지난 1시간 제거)
            try:
                if lines[2].strip() == "True" :
                    self.udfilter_btn_3.setChecked(True)
            except :
                pass
            try:
                if lines[3].strip() == "True" :
                    self.udfilter_btn_4.setChecked(True)
            except :
                pass
            try:
                if lines[4].strip() == "True" :
                    self.udfilter_btn_5.setChecked(True)
            except :
                pass
            try:
                if lines[5].strip() == "True" :
                    self.udfilter_btn_6.setChecked(True)
            except :
                pass

            try:
                self.folder_path_btn.setText(lines[6].strip())
            except:
                self.folder_path_btn.setText('')

            try:
                self.limitcnt_btn.setText(lines[7].strip())
            except:
                self.limitcnt_btn.setText('')
            try:
                self.viewcnt_btn.setText(lines[8].strip())
            except:
                self.viewcnt_btn.setText('')

            try:
                self.member_start_btn.setText(lines[9].strip())
            except:
                self.member_start_btn.setText('')
            try:
                self.member_end_btn.setText(lines[10].strip())
            except:
                self.member_end_btn.setText('')

            try:
                self.delay_start_btn.setText(lines[11].strip())
            except:
                self.delay_start_btn.setText('')
            try:
                self.delay_end_btn.setText(lines[12].strip())
            except:
                self.delay_end_btn.setText('')
            try:
                self.exl_path_btn.setText(lines[13].strip())
                if lines[13].strip() != '' :
                    self.load_excel_data(lines[13].strip())
            except:
                self.exl_path_btn.setText('')

            # 우선순위 필터 로드 (선택안함/관련성/인기도)
            try:
                priority_val = lines[14].strip()
                if priority_val == "0":
                    self.priority_btn_0.setChecked(True)
                elif priority_val == "1":
                    self.priority_btn_1.setChecked(True)
                elif priority_val == "2":
                    self.priority_btn_2.setChecked(True)
                else:
                    self.priority_btn_0.setChecked(True)
            except:
                self.priority_btn_0.setChecked(True)

        if os.path.exists(setting_file_2):

            with open(setting_file_2, 'r') as file:
                lines = file.readlines()
            try:
                self.cnname_except_btn.setPlainText(''.join(lines))
            except:
                self.cnname_except_btn.setPlainText('')

        if os.path.exists(setting_file_3):

            with open(setting_file_3, 'r') as file:
                lines = file.readlines()
            try:
                self.title_except_btn.setPlainText(''.join(lines))
            except:
                self.title_except_btn.setPlainText('')

        # X 버튼 클릭 시 종료 이벤트 처리
        self.setAttribute(Qt.WA_DeleteOnClose)
        self.closeEvent = self.on_closing

    # 프로그램 종료
    def on_closing(self, event):
        try :
            print('종료')
            reply = QMessageBox.question(self, '종료', '프로그램을 종료하시겠습니까?',
                                        QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if reply == QMessageBox.Yes:
                try :
                    self.driver.close()
                except :
                    pass

                event.accept()
            else:
                event.ignore()
        except Exception:
            err = traceback.format_exc()
            self.textEdit_item.append(f"오류가 발생했습니다. {err}")

    def setting_save(self) :

        udfilter_1 = self.udfilter_btn_1.isChecked()
        # udfilter_2 제거됨 (지난 1시간 옵션 삭제)
        udfilter_3 = self.udfilter_btn_3.isChecked()
        udfilter_4 = self.udfilter_btn_4.isChecked()
        udfilter_5 = self.udfilter_btn_5.isChecked()
        udfilter_6 = self.udfilter_btn_6.isChecked()

        folder_path = self.folder_path_btn.text()

        limit_cnt = self.limitcnt_btn.text()
        upper_viewcnt = self.viewcnt_btn.text()

        wishsubcnt_start = self.member_start_btn.text()
        wishsubcnt_end = self.member_end_btn.text()

        rd_time_start = self.delay_start_btn.text()
        rd_time_end = self.delay_end_btn.text()

        exl_path = self.exl_path_btn.text()

        # 우선순위 필터 (선택안함/관련성/인기도)
        if self.priority_btn_0.isChecked():
            priority_filter = "0"
        elif self.priority_btn_1.isChecked():
            priority_filter = "1"
        else:
            priority_filter = "2"

        with open(setting_file_1, 'w') as file:
            file.write(str(udfilter_1) + '\n')
            file.write('False\n')  # 지난 1시간 제거됨 (호환성 유지)
            file.write(str(udfilter_3) + '\n')
            file.write(str(udfilter_4) + '\n')
            file.write(str(udfilter_5) + '\n')
            file.write(str(udfilter_6) + '\n')

            file.write(str(folder_path) + '\n')
            file.write(str(limit_cnt) + '\n')
            file.write(str(upper_viewcnt) + '\n')
            file.write(str(wishsubcnt_start) + '\n')
            file.write(str(wishsubcnt_end) + '\n')
            file.write(str(rd_time_start) + '\n')
            file.write(str(rd_time_end) + '\n')
            file.write(str(exl_path) + '\n')
            file.write(priority_filter)  # 우선순위 필터

        channel_ecp_keywords = self.cnname_except_btn.toPlainText() # 추출 키워드

        with open(setting_file_2, 'w') as file:
            file.write(channel_ecp_keywords) 

        title_ecp_keywords = self.title_except_btn.toPlainText() # 추출 키워드

        with open(setting_file_3, 'w') as file:
            file.write(title_ecp_keywords) 

        self.textEdit_item.appendPlainText(f"현재 설정이 지정되었습니다.")
        QApplication.processEvents()

    # 저장 기본폴더 버튼 이벤트
    def folder_path(self) :
        # Directory 를 선택합니다.
        fname = QFileDialog.getExistingDirectory(self, "Select Directory")
        self.folder_path_btn.setText(fname) 

        self.textEdit_item.appendPlainText(f"저장 기본폴더가 지정되었습니다.")
        QApplication.processEvents()

    # 폴더명 리셋
    def folder_reset(self):
        self.exl_path_btn.clear()

        for folder_name_edit, keyword_text_edit in self.rows:
            folder_name_edit.clear()
            keyword_text_edit.clear()

    # 엑셀 Import
    def exl_load(self):
        fname = QFileDialog.getOpenFileName(self, "File Load", program_directory, 'Excel File (*.xlsx)')

        if fname[0]:
            file_name = fname[0]
            self.exl_path_btn.setText(file_name)

            self.load_excel_data(file_name)

            self.textEdit_item.appendPlainText(f"{file_name}을 불러왔습니다.")
            QApplication.processEvents()

        else:
            self.textEdit_item.appendPlainText("파일을 다시 선택해주세요.")
            QApplication.processEvents()

    # 엑셀 데이타 GUI 화면 반영
    def load_excel_data(self, file_name):
        wb = load_workbook(file_name)
        sheet = wb.active
        
        row_index = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):  # 첫 번째 행(헤더 제외)
            folder_name_value = str(row[0]).strip() if row[0] else ""  # A열 (폴더명)
            keyword_text_value = str(row[1]).strip() if row[1] else ""  # B열 (키워드)

            if row_index < len(self.rows):
                folder_name_edit, keyword_text_edit = self.rows[row_index] 
                folder_name_edit.setText(folder_name_value)  # QLineEdit에 폴더명 입력
                keyword_text_edit.setText(keyword_text_value)  # QTextEdit에 키워드 입력
            else:
                self.add_row()  # 새 행 추가
                folder_name_edit, keyword_text_edit = self.rows[-1]  # 방금 추가된 마지막 행 가져오기
                folder_name_edit.setText(folder_name_value)  # QLineEdit에 폴더명 입력
                keyword_text_edit.setText(keyword_text_value)  # QTextEdit에 키워드 입력
        
            row_index += 1

    # 폴더추가
    def add_row(self):
        row_frame = QFrame()
        row_layout = QHBoxLayout(row_frame)
        
        folder_name = QLineEdit()
        folder_name.setFixedWidth(120)
        folder_name.setFixedHeight(30)
        keyword_text = QTextEdit()

        row_layout.addWidget(folder_name)
        row_layout.addWidget(keyword_text)
        
        self.scroll_layout.addWidget(row_frame)
        self.rows.append((folder_name, keyword_text))

        self.scroll_widget.setLayout(self.scroll_layout)

    # 수집시작
    def main(self):

        # 구독자수 숫자변환
        def convert_subscriber_count(subscriber_str):
            multipliers = {'억': 100000000, '만': 10000, '천': 1000}
            subscriber_str = subscriber_str.replace(' ', '')
            
            for suffix, multiplier in multipliers.items():
                if subscriber_str.endswith(suffix):
                    numeric_part = subscriber_str[:-len(suffix)]
                    return round(float(numeric_part) * multiplier)  # 🔹 `round()` 추가
            
            try:
                return round(float(subscriber_str))  # 🔹 `round()` 추가
            except ValueError:
                return 0

        def clean_for_excel(text):
            """
            엑셀/스프레드시트 저장 시 오류를 유발할 수 있는 제어 문자 및
            비표준 공백 문자를 제거하고, 텍스트를 정리합니다.

            Args:
                text (str): 클리닝할 원본 문자열.

            Returns:
                str: 클리닝된 문자열.
            """
            if not isinstance(text, str):
                return "" # 문자열이 아니면 빈 문자열 반환

            # 1. 널 문자(null byte) 제거: 엑셀에서 가장 흔하게 오류를 유발합니다.
            text = text.replace('\x00', '')

            # 2. 비표준 유니코드 문자 및 제어 문자 제거
            # \x01-\x1F 범위의 제어 문자를 제거합니다. (\t, \n, \r는 제외하고 제거)
            # \t, \n, \r은 엑셀 셀 내 줄바꿈이나 탭으로 사용될 수 있으므로 선택적으로 남겨둡니다.
            # 여기서는 안전하게 '\t', '\n', '\r'만 남기고 나머지는 제거합니다.
            # [^] 안의 문자가 아닌 모든 문자를 찾는 정규식입니다.
            # 참고: 유효한 XML/엑셀 문자 범위만 허용하는 것이 가장 안전합니다.
            # 그러나 여기서는 일반적인 자막 텍스트에 한정하여 처리합니다.
            
            # \t (탭), \n (줄바꿈), \r (캐리지 리턴)을 제외한 모든 ASCII 제어 문자 제거
            # (0x00에서 0x1F 사이)
            text = re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F]', '', text)

            # 3. 비표준 공백 문자 제거 (예: 유니코드 Non-breaking space \xa0)
            text = re.sub(r'[\xa0\u200b\uFEFF]', ' ', text)

            # 4. 마지막으로 텍스트 양 끝의 공백 제거 (선택 사항)
            text = text.strip()

            return text

        # 조회수 숫자변환
        def convert_views(view_str):
            match = re.match(r"([\d\.]+)([천만억]*)회?", view_str)
            
            if not match:
                return None  # 매칭되지 않으면 None 반환
            
            num, unit = match.groups()
            num = float(num)  # 숫자 부분 변환
            
            # 단위별 변환
            unit_multipliers = {"천": 1_000, "만": 10_000, "억": 100_000_000}
            multiplier = unit_multipliers.get(unit, 1)  # 기본값 1 (단위 없음)
            
            return int(num * multiplier)  # 최종 변환 값

        try :

            folder_path = self.folder_path_btn.text()

            limit_cnt = self.limitcnt_btn.text()
            upper_viewcnt = self.viewcnt_btn.text()

            wishsubcnt_start = self.member_start_btn.text()
            wishsubcnt_end = self.member_end_btn.text()

            rd_time_start = self.delay_start_btn.text()
            rd_time_end = self.delay_end_btn.text()

            channel_ecp_keywords = self.cnname_except_btn.toPlainText()
            title_ecp_keywords = self.title_except_btn.toPlainText()

            # 업로드 날짜 필터 (지난 1시간 제거됨)
            udfilter_check_1 = self.udfilter_btn_1.isChecked()  # 선택안함
            udfilter_check_3 = self.udfilter_btn_3.isChecked()  # 오늘
            udfilter_check_4 = self.udfilter_btn_4.isChecked()  # 이번 주
            udfilter_check_5 = self.udfilter_btn_5.isChecked()  # 이번 달
            udfilter_check_6 = self.udfilter_btn_6.isChecked()  # 올해

            # 우선순위 필터 (선택안함/관련성/인기도)
            priority_check_0 = self.priority_btn_0.isChecked()  # 선택안함
            priority_check_1 = self.priority_btn_1.isChecked()  # 관련성
            priority_check_2 = self.priority_btn_2.isChecked()  # 인기도

            ud_text = ''
            if udfilter_check_3 :
                ud_text = '오늘'
            if udfilter_check_4 :
                ud_text = '이번 주'
            if udfilter_check_5 :
                ud_text = '이번 달'
            if udfilter_check_6 :
                ud_text = '올해'

            # 우선순위 텍스트 설정
            if priority_check_1:
                priority_text = '관련성'
            elif priority_check_2:
                priority_text = '인기도'
            else:
                priority_text = ''


            # 유효성 검사

            profile_path = r"C:\selenium_profile\vidiq"

            if not os.path.exists(profile_path):
                QMessageBox.information(self, "설정오류", "본 프로그램 종료 후 ShortsSetting_ver1.0 부터 진행 후 다시 실행하세요.")
                QApplication.processEvents()
                return 0
            

            if folder_path == '' :
                QMessageBox.information(self, "설정오류", "저장 기본폴더를 입력하세요.")
                QApplication.processEvents()
                return 0

            try :
                limit_cnt = int(limit_cnt)
            except :
                QMessageBox.information(self, "설정오류", "키워드당 수집개수를 숫자로 입력하세요.")
                QApplication.processEvents()
                return 0

            try :
                upper_viewcnt = int(upper_viewcnt)
            except :
                QMessageBox.information(self, "설정오류", "조회수(이상)를 숫자로 입력하세요.")
                QApplication.processEvents()
                return 0    

            try :
                wishsubcnt_start = int(wishsubcnt_start)
            except :
                QMessageBox.information(self, "설정오류", "채널구독자수(명) 시작 값을 숫자로 입력하세요.")
                QApplication.processEvents()
                return 0 
            try :
                wishsubcnt_end = int(wishsubcnt_end)
            except :
                QMessageBox.information(self, "설정오류", "채널구독자수(명) 종료 값을 숫자로 입력하세요.")
                QApplication.processEvents()
                return 0 
            if wishsubcnt_start > wishsubcnt_end :  
                QMessageBox.information(self, "설정오류", "채널구독자수(명) 시작값을 종료값 보다 작게 입력하세요.")
                QApplication.processEvents()
                return 0

            try :
                rd_time_start = int(rd_time_start)
            except :
                QMessageBox.information(self, "설정오류", "딜레이(초) 시작 값을 숫자로 입력하세요.")
                QApplication.processEvents()
                return 0 
            try :
                rd_time_end = int(rd_time_end)
            except :
                QMessageBox.information(self, "설정오류", "딜레이(초) 종료 값을 숫자로 입력하세요.")
                QApplication.processEvents()
                return 0
            if rd_time_start > rd_time_end :  
                QMessageBox.information(self, "설정오류", "딜레이(초) 시작값을 종료값 보다 작게 입력하세요.")
                QApplication.processEvents()
                return 0


            folder_lines = []
            
            for folder_name, keyword_text in self.rows:
                folder_value = folder_name.text().strip()
                keyword_lines = [line.strip() for line in keyword_text.toPlainText().split("\n") if line.strip()]  # 공란 제거

                if folder_value == '' :
                    break
                
                if keyword_lines:  # keyword_text가 비어 있지 않은 경우만 추가
                    folder_lines.append([folder_value] + keyword_lines)

            if len(folder_lines) == 0 :
                QMessageBox.information(self, "설정오류", "수집진행할 폴더명/키워드명이 존재하지 않습니다.(키워드명 입력시 폴더명 필수값)")
                QApplication.processEvents()
                return 0

            now_time = datetime.today().strftime("%Y-%m-%d_%H:%M")
            self.textEdit_item.appendPlainText(f'\n▶▶▶ 유튜브 정보수집을 시작합니다.({now_time})\n')
            QApplication.processEvents()

            # ChromeOptions 설정
            profile_path = r"C:\selenium_profile\vidiq"
            options = webdriver.ChromeOptions()
            options.add_argument(fr"--user-data-dir={profile_path}")
            options.add_argument(r'--profile-directory=Profile 1')
            # options.add_argument(r"--user-data-dir=C:\Users\james\AppData\Local\Google\Chrome\User Data")  # 본인 PC 계정으로 변경
            # options.add_argument(r"C:\Users\james\AppData\Local\Google\Chrome\User Data")  # 본인 PC 계정으로 변경
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_argument('Accept-Language=ko-KR,ko;q=0.8,en-US;q=0.5,en;q=0.3')
            options.add_argument("--no-sandbox")
            options.add_argument("--disable-gpu")
            options.add_argument("--disable-page-load-metrics")
            options.add_argument("--disable-devtools-experiments")
            options.add_argument("--disable-geolocation")
            options.add_experimental_option("useAutomationExtension", False)
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_argument("--disable-dev-shm-usage")
            options.add_argument(f"--force-device-scale-factor=0.7")
            options.add_experimental_option("prefs", {
                "credentials_enable_service": False, 
                "profile.password_manager_enabled": False  
            })

            # options.add_argument(f"--load-extension={vidiq_latest_folder}")

            driver = webdriver.Chrome(options=options)

            driver.execute_script("window.open('about:blank');")
            time.sleep(1)

            # 수집일자 폴더 세팅
            current_time = datetime.now().strftime("%y%m%d")
            date_folder = folder_path + f'\\{current_time}'

            if not os.path.exists(date_folder):
                os.makedirs(date_folder)

            channel_ecp_keywords = [line.strip() for line in channel_ecp_keywords.split("\n") if line.strip()]
            title_ecp_keywords = [line.strip() for line in title_ecp_keywords.split("\n") if line.strip()]

            # 엑셀세팅
            save_date = datetime.today().strftime("%Y%m%d_%H%M%S")
            exlsave_name = date_folder + f"\\{save_date}.xlsx"

            wb = load_workbook(exl_sample)
            sheet = wb.active
            last_row = sheet.max_row

            search_num = 1

            except_link = []

            # 반복문으로 수집시작
            for folder_line in folder_lines :

                folder_name = folder_line[0]
                self.searchfolder_btn.setText(folder_name)
                QApplication.processEvents()

                for keyword in folder_line[1:] :

                    kwd_search_num = 0

                    self.searchkwd_btn.setText(keyword)
                    QApplication.processEvents()

                    self.textEdit_item.appendPlainText(f'\n▶▶ {keyword}(폴더명 : {folder_name}) 확인 중\n')
                    QApplication.processEvents()

                    driver.switch_to.window(driver.window_handles[0])
                    time.sleep(1)

                    query = urllib.parse.quote(keyword) 

                    handles = driver.window_handles
                    driver.switch_to.window(handles[0])

                    link = f'https://www.youtube.com/results?search_query={query}'
                    driver.get(link)
                    time.sleep(5)

                    # 유튜브 검색필터 추가 (새로운 UI 구조에 맞게 수정)
                    # 필터 순서: 업로드날짜(8-11), 우선순위(23-24)

                    filter_steps = []

                    # 1. 업로드 날짜 필터 (선택안함이면 스킵)
                    if udfilter_check_1 == False:
                        filter_steps.append({'search_txt': ud_text, 'range': (8, 12)})

                    # 2. 우선순위 필터 (선택안함이면 스킵)
                    if priority_check_0 == False:
                        filter_steps.append({'search_txt': priority_text, 'range': (23, 25)})

                    for step in filter_steps:
                        search_txt = step['search_txt']
                        filter_range = step['range']

                        filter_c_tag = driver.find_element(By.CSS_SELECTOR, ".yt-spec-button-shape-next.yt-spec-button-shape-next--text.yt-spec-button-shape-next--mono.yt-spec-button-shape-next--size-m.yt-spec-button-shape-next--icon-trailing.yt-spec-button-shape-next--enable-backdrop-filter-experiment")
                        filter_c_tag.click()
                        time.sleep(2)

                        filter_tags = driver.find_elements(By.CSS_SELECTOR, "ytd-search-filter-renderer")
                        rst_filter_tags = filter_tags[filter_range[0]:filter_range[1]]

                        for rst_filter_tag in rst_filter_tags:
                            current_tag = rst_filter_tag.find_element(By.CSS_SELECTOR, "#label")
                            current_text = current_tag.text.strip()

                            if current_text == search_txt:
                                current_tag.click()
                                time.sleep(3)
                                break

                    while 1 :

                        driver.switch_to.window(driver.window_handles[0])
                        time.sleep(1)

                        # 현재 페이지 높이 저장
                        last_height = driver.execute_script("return document.documentElement.scrollHeight")

                        soup = BeautifulSoup(driver.page_source, "html.parser")
                        videos = soup.select("ytd-video-renderer")

                        video_links = []

                        for idx, video in enumerate(videos,start=1) :

                            descriptions = video.select_one("#metadata-line").text.split("\n")

                            # 제목
                            title_tag = video.select_one("#video-title")
                            title = title_tag.text.strip()

                            # 링크
                            links = title_tag.get("href")
                            links = links.split('&')
                            link = 'https://www.youtube.com' + links[0]

                            if 'shorts' not in link :
                                continue

                            if link in except_link :
                                continue

                            except_link.append(link)

                            # 제목 금칙어 해당여부
                            title_pass = None
                            for title_ecp_keyword in title_ecp_keywords :
                                if title_ecp_keyword in title :
                                    title_pass = True
                                    break
                            if title_pass == True :
                                print(f'▷ {title} : 영상제목 금지어 포함되어 수집제외({title_ecp_keyword})[1차 필터링]')
                                self.textEdit_item.appendPlainText(f'▷ {title} : 영상제목 금지어 포함되어 수집제외({title_ecp_keyword})[1차 필터링]')
                                QApplication.processEvents()
                                continue            

                            # 채널명
                            channel_tags = video.select("a.yt-simple-endpoint.style-scope.yt-formatted-string")
                            channel = channel_tags[1].text.strip()

                            # 채널명 금칙어 해당여부
                            cnname_pass = None
                            for channel_ecp_keyword in channel_ecp_keywords  :
                                if channel_ecp_keyword in channel :
                                    cnname_pass = True
                                    break

                            if cnname_pass == True :
                                print(f'▷ {title} : 채널명 금지어 포함되어 수집제외({channel_ecp_keyword})[1차 필터링]')
                                self.textEdit_item.appendPlainText(f'▷ {title} : 채널명 금지어 포함되어 수집제외({channel_ecp_keyword})[1차 필터링]')
                                QApplication.processEvents()
                                continue 

                            # 조회수
                            if ' 없음' not in descriptions[3] :
                                view_cnts = descriptions[3].replace('조회수 ','')
                                view_cnt = convert_views(view_cnts)
                                print('view_cnt', view_cnt)
                            else :
                                view_cnt = 0

                            if view_cnt < upper_viewcnt : # 조회수 기준 미달시 패스
                                print(f'▷ {title} : 조회수 미충족 수집제외({view_cnt}회)[1차 필터링]')
                                self.textEdit_item.appendPlainText(f'▷ {title} : 조회수 미충족 수집제외({view_cnt}회)[1차 필터링]')
                                QApplication.processEvents()
                                continue


                            # 스트리밍 여부
                            upload_date = descriptions[4]
                            if '스트리밍' in upload_date :
                                print(f'▷ {title} : 스트리밍 수집제외[1차 필터링]')
                                self.textEdit_item.appendPlainText(f'▷ {title} : 스트리밍 수집제외[1차 필터링]')
                                QApplication.processEvents()
                                continue

                            video_links.append([title, link, channel])
                            print(f'□ {title} : 1차 링크 수집완료')

                        if len(video_links) > 0 : # 조회수 추가

                            driver.switch_to.window(driver.window_handles[1])
                            time.sleep(1)

                            for video_link in video_links :

                                try :

                                    ytb_title = video_link[0]
                                    ytb_link = video_link[1]
                                    ytb_channel = video_link[2]

                                    print()
                                    print('================================================')
                                    print(ytb_link)
                                    print(ytb_title)

                                    driver.get(ytb_link)
                                    time.sleep(5)

                                    soup = BeautifulSoup(driver.page_source, "html.parser")

                                    # 영상 정보
                                    def normalize_number(text):
                                        """
                                        '1.5만', '1,841,537'와 같은 텍스트를 정수(Integer)로 변환합니다.
                                        """
                                        if not text:
                                            return 0
                                        
                                        # 숫자와 '.'을 제외한 모든 문자 제거
                                        cleaned_text = re.sub(r'[^\d.]', '', text)
                                        
                                        # '만' 단위 처리 (JSON 데이터에서는 숫자로 치환된 경우가 많지만, 혹시 모를 경우 대비)
                                        if '만' in text:
                                            try:
                                                return int(float(cleaned_text) * 10000)
                                            except ValueError:
                                                pass

                                        # 일반적인 콤마 제거 및 정수 변환
                                        try:
                                            return int(cleaned_text.replace(',', '').replace('.', ''))
                                        except ValueError:
                                            return 0

                                    def extract_video_stats(data):
                                        """
                                        파싱된 JSON 데이터(Python Dict)에서 조회수, 좋아요 수, 댓글 수를 추출합니다.
                                        (이전 답변에서 사용된 경로 기반)
                                        """
                                        stats = {
                                            "view_count": 0,
                                            "like_count": 0,
                                            "comment_count": 0
                                        }
                                        

                                        # 1. 좋아요 수 추출 (⭐ 이 부분을 수정했습니다.)
                                        try:
                                            # 현재 화면에 표시되는 좋아요 수는 'defaultButtonViewModel'의 'title' 필드에 있습니다.
                                            like_view_model_default = data['overlay']['reelPlayerOverlayRenderer']['buttonBar']['reelActionBarViewModel']['buttonViewModels'][0]['likeButtonViewModel']['toggleButtonViewModel']['toggleButtonViewModel']['defaultButtonViewModel']
                                            
                                            # 'title'은 '94'와 같은 단순한 문자열입니다.
                                            like_count_text = like_view_model_default['buttonViewModel']['title']
                                            stats['like_count'] = normalize_number(like_count_text)

                                        except (KeyError, IndexError, TypeError, AttributeError):
                                            # 좋아요 수 추출 실패 시 'factoid'에서 추출 시도 (설명 패널)
                                            try:
                                                factoids = data['engagementPanels'][1]['engagementPanelSectionListRenderer']['content']['structuredDescriptionContentRenderer']['items'][0]['videoDescriptionHeaderRenderer']['factoid']
                                                for factoid in factoids:
                                                    if factoid.get('factoidRenderer', {}).get('label', {}).get('simpleText') == '좋아요':
                                                        like_count_text = factoid['factoidRenderer']['value']['simpleText']
                                                        stats['like_count'] = normalize_number(like_count_text)
                                                        break
                                            except (KeyError, IndexError, TypeError, AttributeError):
                                                pass # 추출 실패 시 0 유지

                                        # 2. 댓글 수 추출
                                        try:
                                            # 댓글 버튼의 'title' 필드에서 숫자 추출
                                            comment_button = data['overlay']['reelPlayerOverlayRenderer']['buttonBar']['reelActionBarViewModel']['buttonViewModels'][2]['buttonViewModel']
                                            comment_count_text = comment_button['title']
                                            stats['comment_count'] = normalize_number(comment_count_text)

                                        except (KeyError, IndexError, TypeError):
                                            # 대안: 댓글 패널 헤더에서 추출 시도
                                            try:
                                                comment_header = data['engagementPanels'][0]['engagementPanelSectionListRenderer']['header']['engagementPanelTitleHeaderRenderer']
                                                comment_count_text = comment_header['contextualInfo']['runs'][0]['text']
                                                stats['comment_count'] = normalize_number(comment_count_text)
                                            except (KeyError, IndexError, TypeError):
                                                pass # 추출 실패 시 0 유지

                                        # 3. 조회수 추출
                                        try:
                                            # 설명 패널의 videoDescriptionHeaderRenderer 내 views 필드에서 추출
                                            views_text = data['engagementPanels'][1]['engagementPanelSectionListRenderer']['content']['structuredDescriptionContentRenderer']['items'][0]['videoDescriptionHeaderRenderer']['views']['simpleText']
                                            
                                            match = re.search(r'(\d[\d,.]*)', views_text)
                                            if match:
                                                stats['view_count'] = normalize_number(match.group(1))

                                        except (KeyError, IndexError, TypeError):
                                            pass # 추출 실패 시 0 유지

                                        return stats

                                    yt_data_dict = driver.execute_script("return window.ytInitialData;")

                                    extracted_stats = extract_video_stats(yt_data_dict)

                                    view_count = extracted_stats['view_count']
                                    like_count = extracted_stats['like_count']
                                    comment_count = extracted_stats['comment_count']

                                    if view_count < upper_viewcnt :
                                        print(f'▷ {ytb_title} : 조회수 미충족 수집제외({view_count}회)[2차 필터링]')
                                        self.textEdit_item.appendPlainText(f'▷ {ytb_title} : 조회수 미충족 수집제외({view_count}회)[2차 필터링]')
                                        QApplication.processEvents()
                                        continue

                                    like_count = extracted_stats['like_count']
                                    comment_count = extracted_stats['comment_count']

                                    # 구독자수

                                    subscriber_p_tag = soup.find(
                                        'p', 
                                        string=lambda text: text and 'Subscribers' in text
                                    )

                                    if subscriber_p_tag:
                                        # 텍스트 추출 (앞뒤 공백 제거)
                                        full_subscriber_text = subscriber_p_tag.get_text(strip=True)

                                        # 2. 텍스트에서 숫자 부분만 분리하여 추출 (옵션)
                                        # 정규표현식을 사용하여 숫자와 '만', '억' 등의 단위(한글 포함)만 추출합니다.
                                        match = re.search(r'([\d.,가-힣]+)\s*Subscribers', full_subscriber_text)
                                        
                                        if match:
                                            subscriber_count = match.group(1).strip()
                                            subscribe_cnt = convert_subscriber_count(subscriber_count)
                                        else:
                                            subscribe_cnt = 0
                                            
                                    else:
                                        subscribe_cnt = 0

                                    

                                    if not (wishsubcnt_start <= subscribe_cnt <= wishsubcnt_end) :
                                        print(f'▷ {ytb_title} : 구독자수 기준 미충족 수집제외[2차 필터링]')
                                        self.textEdit_item.appendPlainText(f'▷ {ytb_title} : 구독자수 미충족 수집제외({subscribe_cnt}명)[2차 필터링]')
                                        QApplication.processEvents()
                                        continue


                                    target_text = "Views per hour"
                                    target_span = soup.find('span', string=target_text)

                                    extracted_number = None

                                    if target_span:

                                        number_span = target_span.find_next_sibling('span')
                                        
                                        if number_span:
                                            # 4. 다음 <span> 태그의 텍스트를 추출하고 앞뒤 공백을 제거합니다.
                                            viewsper_hour = number_span.get_text(strip=True)
                                            
                                    else:
                                        viewsper_hour = 0

                                    # --- 최종 결과 출력 ---

                                    target_text = "Engagement"
                                    target_span = soup.find('span', string=target_text)

                                    extracted_value = None

                                    if target_span:
                                        
                                        # 3. 찾은 태그의 바로 다음 형제(sibling) 태그를 찾습니다.
                                        # find_next_sibling('span')을 사용하여 바로 다음의 <span> 태그를 검색합니다.
                                        value_span = target_span.find_next_sibling('span')
                                        
                                        if value_span:
                                            # 4. 다음 <span> 태그의 텍스트를 추출하고 앞뒤 공백을 제거합니다.
                                            engagement = value_span.get_text(strip=True)
                                        else:
                                            engagement = '-'
                                            
                                    else:
                                        engagement = '-'

                                    ############################################# 자막

                                    video_id = ytb_link.split("/shorts/")[1].split("?")[0]
                                    script_url = f"https://www.youtube.com/watch?v={video_id}"

                                    ydl_opts = {
                                        "quiet": True,
                                        "no_warnings": True,
                                        "ignoreerrors": True,
                                        "skip_download": True,
                                        "http_headers": {
                                            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36"
                                        }
                                    }

                                    script_text = ""

                                    try:
                                        with YoutubeDL(ydl_opts) as ydl:
                                            info = ydl.extract_info(script_url, download=False)

                                            captions = info.get("subtitles") or info.get("automatic_captions")

                                            caption_url = None

                                            if captions:
                                                for lang in ["ko", "ko-KR", "en"]:
                                                    if lang in captions:
                                                        caption_url = captions[lang][0]["url"]
                                                        break

                                            if not caption_url:
                                                script_text = "자막 정보 없음"

                                            else:
                                                headers = {
                                                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120 Safari/537.36"
                                                }
                                                res = requests.get(caption_url, headers=headers)

                                                if res.status_code != 200:
                                                    script_text = "자막 정보 없음"
                                                else:
                                                    raw_data = res.text

                                                    if raw_data.lstrip().startswith("<html"):
                                                        script_text = "자막 정보 없음"

                                                    # JSON
                                                    elif raw_data.lstrip().startswith("{"):
                                                        data = json.loads(raw_data)
                                                        all_lines = []
                                                        for event in data.get("events", []):
                                                            if "segs" in event:
                                                                for seg in event["segs"]:
                                                                    text_piece = seg.get("utf8", "").strip()
                                                                    if text_piece:
                                                                        all_lines.append(text_piece)
                                                        script_text = " ".join(all_lines)

                                                    # VTT
                                                    else:
                                                        lines = []
                                                        for line in raw_data.split("\n"):
                                                            t = line.strip()
                                                            if (
                                                                t
                                                                and "-->" not in t
                                                                and not t.startswith("WEBVTT")
                                                                and not t.isdigit()
                                                            ):
                                                                lines.append(t)
                                                        script_text = " ".join(lines)

                                    except Exception as e:
                                        print("오류:", e)


                                    ############################################# 자막



                                    button = driver.execute_script("""
                                        const btn = document.querySelector('yt-shorts-suggested-action-view-model button');
                                        if (btn) { btn.click(); return true; }
                                        return false;
                                    """)

                                    if button:
                                        print("제품 보기 버튼 클릭 성공")
                                    else:
                                        print("버튼을 찾지 못했습니다")
                                        continue


                                    # 2. 쇼핑 패널 로드 대기
                                    product_list_xpath = '//ytd-product-list-renderer[contains(@panel-target-id, "shopping_panel")]'
                                    WebDriverWait(driver, 15).until(
                                        EC.presence_of_element_located((By.XPATH, product_list_xpath))
                                    )

                                    # 3. 모든 상품 항목 찾기
                                    product_items_xpath = '//ytd-product-list-item-renderer'
                                    product_items = driver.find_elements(By.XPATH, product_items_xpath)

                                    if len(product_items) == 0 :
                                        print(f"📦 0개 제품 패스")
                                        continue


                                    extracted_urls = []

                                    for i, item in enumerate(product_items):

                                        # 3개 이상 수집되면 즉시 종료
                                        if len(extracted_urls) >= 3:
                                            break

                                        try:
                                            # 현재 탭 핸들 저장
                                            current_handle = driver.current_window_handle
                                            before = len(driver.window_handles)

                                            # '자세히 보기' 버튼
                                            view_button = item.find_element(
                                                By.XPATH, ".//button[contains(@aria-label, '자세히 보기')]"
                                            )

                                            # 클릭 보조 (스크롤 + JS 클릭)
                                            driver.execute_script("arguments[0].scrollIntoView(true);", view_button)
                                            driver.execute_script("arguments[0].click();", view_button)

                                            # 새 탭 열릴 때까지 기다림
                                            WebDriverWait(driver, 10).until(
                                                lambda d: len(d.window_handles) == before + 1
                                            )

                                            # 마지막 탭으로 이동
                                            new_tab = driver.window_handles[-1]
                                            driver.switch_to.window(new_tab)

                                            time.sleep(3)
                                            final_url = driver.current_url
                                            extracted_urls.append(final_url)

                                            # 새 탭 닫기 후 원래 탭으로 복귀
                                            driver.close()
                                            driver.switch_to.window(current_handle)

                                            # 쇼핑 패널 재로딩 대기
                                            WebDriverWait(driver, 10).until(
                                                EC.presence_of_element_located((By.XPATH, product_list_xpath))
                                            )

                                        except Exception as e:
                                            try:
                                                driver.switch_to.window(current_handle)
                                            except:
                                                pass
                                            continue

                                    # 엑셀저장

                                    sheet[f'A{search_num + last_row}'] = folder_name # 폴더명
                                    sheet[f'B{search_num + last_row}'] = keyword # 키워드
                                    sheet[f'C{search_num + last_row}'] = ytb_title # 영상제목

                                    exl_link = ytb_link.replace('?v=','/')
                                    sheet[f'D{search_num + last_row}'] = exl_link # 링크
                                    sheet[f'D{search_num + last_row}'].hyperlink = sheet[f'D{search_num + last_row}'].value
                                    sheet[f'D{search_num + last_row}'].style = "Hyperlink"


                                    sheet[f'E{search_num + last_row}'] = view_count # 조회수
                                    sheet[f'F{search_num + last_row}'] = subscribe_cnt # 구독자수
                                    sheet[f'G{search_num + last_row}'] = int(comment_count) # 댓글수
                                    sheet[f'H{search_num + last_row}'] = int(like_count)  # 좋아요 수

                                    sheet[f'I{search_num + last_row}'] = viewsper_hour  # viewsper_hour

                                    sheet[f'J{search_num + last_row}'] = engagement  # engagement

                                    sheet[f'E{search_num + last_row}'].number_format = '#,##0'
                                    sheet[f'F{search_num + last_row}'].number_format = '#,##0'
                                    sheet[f'G{search_num + last_row}'].number_format = '#,##0'
                                    sheet[f'H{search_num + last_row}'].number_format = '#,##0'
                                    sheet[f'I{search_num + last_row}'].number_format = '#,##0'


                                    try :
                                        sheet[f'K{search_num + last_row}'] = str(script_text)  # engagement
                                    except :
                                        sheet[f'K{search_num + last_row}'] = str(clean_for_excel(script_text))
                                    # L, M, N 열에 extracted_urls 입력
                                    target_row = search_num + last_row

                                    columns = ['L', 'M', 'N']

                                    for idx, col in enumerate(columns):
                                        try:
                                            sheet[f'{col}{target_row}'] = extracted_urls[idx]
                                            sheet[f'{col}{target_row}'].hyperlink = extracted_urls[idx]
                                            sheet[f'{col}{target_row}'].style = "Hyperlink"
                                        except IndexError:
                                            # extracted_urls 개수가 부족하면 비워둠
                                            sheet[f'{col}{target_row}'] = ""


                                    wb.save(exlsave_name)

                                    search_num += 1
                                    kwd_search_num +=1

                                    print(f'▶ {kwd_search_num}/{limit_cnt} - {ytb_title} : 수집완료')
                                    self.textEdit_item.appendPlainText(f'▶ {kwd_search_num}/{limit_cnt} - {ytb_title} : 수집완료')
                                    QApplication.processEvents()

                                    if kwd_search_num == limit_cnt :
                                        break

                                    time.sleep(random.uniform(rd_time_start, rd_time_end))

                                except Exception:
                                    err = traceback.format_exc()
                                    print(err)
                                    pass 


                        driver.switch_to.window(driver.window_handles[0])
                        time.sleep(1)

                        if kwd_search_num == limit_cnt :
                            break

                        # 페이지 끝까지 스크롤
                        driver.find_element(By.TAG_NAME, "body").send_keys(Keys.END)

                        # 로딩 대기
                        time.sleep(5)
                        

                        # 새로운 페이지 높이 가져오기
                        new_height = driver.execute_script("return document.documentElement.scrollHeight")

                        # 더 이상 스크롤할 곳이 없으면 종료
                        if new_height == last_height:
                            break

                    time.sleep(1)

            now_time = datetime.today().strftime("%Y-%m-%d_%H:%M")
            self.textEdit_item.appendPlainText(f'\n▶▶▶ 유튜브 정보수집이 완료되었습니다.({now_time})\n')
            QApplication.processEvents()

            QMessageBox.information(self, "수집완료", "유튜브 정보수집이 완료되었습니다.")
            QApplication.processEvents()

        except Exception:
            err = traceback.format_exc()
            self.textEdit_item.appendPlainText(err) 
            now_time = datetime.today().strftime("%Y-%m-%d_%H:%M")
            self.textEdit_item.appendPlainText(f'\n▶▶▶ 오류가 발생되어 정보수집이 중지되었습니다.({now_time})\n')
            QApplication.processEvents()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainDialog()
    window.show()
    sys.exit(app.exec_())